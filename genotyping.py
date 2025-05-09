###Calculates Chi^2 and Odd Ratio for genotyping data###

#Takes an excel in ./input, reads genotyping data from all columns except for the last and groups from the last column
#Then calculates Chi^2 and Odd Ratio for genotypes, dominance and alleles 

##Functions
def critical_error(message):
    logging.critical(message)
    logging.info("Press enter to Quit")
    input()
    exit()

def chi(df):
    chi2_result = stats.chi2_contingency(df)
    chi2_p = chi2_result.pvalue
    chi2_df = pd.DataFrame(chi2_result.expected_freq, index = df.index, columns = df.columns.values)
    return chi2_p, chi2_df

def oddratio(df):   
    res = sm.stats.Table2x2(np.array(df))
    OR = getattr(res,"oddsratio")
    LCB = res.oddsratio_confint()[0]
    UCB = res.oddsratio_confint()[1]
    p_value = res.oddsratio_pvalue()
    return [OR, LCB, UCB, p_value]
    
def chi_writer (df, r, c, skip2 = True, header = True):
    rows = dataframe_to_rows(df)
    for r_idx, row in enumerate(rows, 1):
        #Row 2 just has the SNP name
        if skip2:
            if r_idx == 2:
                continue
            if r_idx > 2:
                r_idx -= 1
        
        #Skips writing the column names
        if not header:
            if r_idx == 1:
                continue
        
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row = r_idx + r, column = c_idx + c, value = value)
    return r_idx, c_idx

def OR_writer(df, r, c, header = True):
    OR_header = ["Odd ratio", "CI (95%) lower bound", "CI (95%) upper bound", "P-value"]
    for el in range(0, 4):
        if header:
            sheet.cell(row = r, column = c + el, value = OR_header[el])
            r += 1
        sheet.cell(row = r, column = c + el, value = df[el])
        sheet.cell(row = r, column = c + el).alignment = centered_v
        sheet.merge_cells(start_row = r, start_column = c + el, end_row = r + 1, end_column = c + el)
        if header:
            r -= 1

##Imports
try:
    import logging
    logging.basicConfig(level = logging.INFO, format = "%(asctime)s.%(msecs)03d %(levelname)s: %(message)s",datefmt = "%H:%M:%S")
    import sys
    import os
    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows
    import pandas as pd
    import numpy as np
    import scipy.stats as stats
    import statsmodels.api as sm
except ModuleNotFoundError:
    critical_error("Module not found, please make sure all required modules are installed")

#Change working directory to the python file location
os.chdir(sys.path[0])

##Styles
bold = openpyxl.styles.Font(bold = True)
centered_v = openpyxl.styles.Alignment(vertical = "center")
red = openpyxl.styles.Font(color = "FF0000")

##Code

#Read first file in "./input" and the first sheet in that file and save it as dataframe
try:
    input = os.listdir("./input")[0]
    df = pd.read_excel(f"./input/{input}")
    logging.info(f"Opening: {input}")
except (IndexError, ValueError):
    if not os.path.exists("./input"):
        os.mkdir("./input")
    critical_error("No excel file in input")

#Get group data from last column and genotyping data from all other columns
genotypes = df.columns.values[:-1]
group = df.columns.values[-1]

#Initialize output file to write to and removes worksheet created by default
output = openpyxl.Workbook()
output.remove(output["Sheet"])

#Calculate all chi2 and odds ratio statistics
for snp in genotypes:
    logging.info(f"Working on {snp}")
    
    #Initialize dataframe to work with, ignoring NA and spaces. Sorts so heterozygotes are in the middle
    initial_df = pd.crosstab(df[snp], df[group], dropna = True)
    initial_df = initial_df.sort_values(snp)

    #Initialize file and sheet to write to
    output.create_sheet(title = snp)
    sheet = output[snp]

    #Genotyping data is only going to be 2x2 or 3x2, process 2x2 data
    if initial_df.shape == (2, 2):
        chi2_2x2_p, chi2_2x2_df = chi(initial_df)
        OR_2x2 = oddratio(initial_df)

        sheet["A1"] = snp
        sheet["A3"] = "Genotype"
        chi_writer(initial_df, r = 2, c = 0)
        sheet["D3"] = "P-value"
        sheet["D4"] = chi2_2x2_p
        if chi2_2x2_p <= 0.05:
            sheet["D4"].font = red
        sheet["D4"].alignment = centered_v
        sheet.merge_cells("D4:D5")
        OR_writer(OR_2x2, r = 3, c = 5)
        sheet["J3"] = "Expected"
        chi_writer(chi2_2x2_df, r = 2, c = 9)
        
    
    #Test if dataframe is 3x2, otherwise skip
    if initial_df.shape != (3, 2):
        logging.warning("Data shape is not 2x2 or 3x2. Skipping")
        sheet["A1"] = snp
        sheet["A3"] = "Data shape is not 2x2 or 3x2"
        continue
    
    #Chi square on 3x2
    logging.info("Calculating chi square on 3x2")
    chi2_3x2_p, chi2_3x2_df = chi(initial_df)
    
    #Dominance 2x2 (adding heterozygotes to either homozygote sequentially) chi square and odd ratio
    logging.info("Calculating chi square and odd ratio on dominance 2x2")
    idx = [i for i in initial_df.index]
    
    first_line = initial_df.loc[idx[0]] + initial_df.loc[idx[1]]
    second_line = initial_df.loc[idx[2]]
    dominance_df1 = pd.DataFrame(data = [first_line, second_line], index = [f"{idx[0]} + {idx[1]}", idx[2]])
    chi2_2x2_1_p, chi2_2x2_1_df = chi(dominance_df1)
    OR_2x2_1 = oddratio(dominance_df1)
    
    first_line = initial_df.loc[idx[0]]
    second_line = initial_df.loc[idx[1]] + initial_df.loc[idx[2]]
    dominance_df2 = pd.DataFrame(data = [first_line, second_line], index = [idx[0], f"{idx[1]} + {idx[2]}"])
    chi2_2x2_2_p, chi2_2x2_2_df = chi(dominance_df2)  
    OR_2x2_2 = oddratio(dominance_df2)
    
    #Alleles 2x2 (adding heterozygotes to both homozygotes doubled) chi square and odd ratio
    logging.info("Calculating chi square and odd ratio on dominance 2x2")
    first_line = initial_df.loc[idx[0]] + initial_df.loc[idx[0]] + initial_df.loc[idx[1]]
    second_line = initial_df.loc[idx[2]] + initial_df.loc[idx[2]] + initial_df.loc[idx[1]]
    allele_df = pd.DataFrame(data = [first_line, second_line], index = [f"{idx[0][0]}", f"{idx[2][0]}"])
    chi2_2x2_a_p, chi2_2x2_a_df = chi(allele_df)
    OR_2x2_a = oddratio(allele_df)
    
    #Write 3x2
    sheet["A1"] = snp
    sheet["A3"] = "Genotypes"
    sheet["A3"].font = bold
    sheet["A5"] = "Genotype"
    chi_writer(initial_df, r = 4, c = 0)
    sheet["D5"] = "P-value"
    sheet["D6"] = chi2_3x2_p
    if chi2_3x2_p <= 0.05:
        sheet["D6"].font = red
    sheet["D6"].alignment = centered_v
    sheet.merge_cells("D6:D8")
    sheet["F5"] = "Expected"
    chi_writer(chi2_3x2_df, r = 4, c = 5)
       
    #Write Dominance
    sheet["A10"] = "Dominance"
    sheet["A10"].font = bold
    sheet["A12"] = "Genotype"
    chi_writer(dominance_df1, r = 11, c = 0)
    sheet["D12"] = "P-value"
    sheet["D13"] = chi2_2x2_1_p
    if chi2_2x2_1_p <= 0.05:
        sheet["D13"].font = red
    sheet["D13"].alignment = centered_v
    sheet.merge_cells("D13:D14")
    OR_writer(OR_2x2_1, r = 12, c = 5)
    sheet["J12"] = "Expected"
    chi_writer(chi2_2x2_1_df, r = 11, c = 9)

    chi_writer(dominance_df2, r = 13, c = 0, header = False)
    sheet["D15"] = chi2_2x2_2_p
    if chi2_2x2_2_p <= 0.05:
        sheet["D15"].font = red
    sheet["D15"].alignment = centered_v
    sheet.merge_cells("D15:D16")
    OR_writer(OR_2x2_2, r = 15, c = 5, header = False)
    chi_writer(chi2_2x2_2_df, r = 13, c = 9, header = False)
       
    #Write Alleles
    sheet["A18"] = "Alleles"
    sheet["A18"].font = bold
    sheet["A20"] = "Allele"
    chi_writer(allele_df, r = 19, c = 0)
    sheet["D20"] = "P-value"
    sheet["D21"] = chi2_2x2_a_p
    if chi2_2x2_a_p <= 0.05:
        sheet["D21"].font = red
    sheet["D21"].alignment = centered_v
    sheet.merge_cells("D21:D22")
    OR_writer(OR_2x2_a, r = 20, c = 5)
    sheet["J20"] = "Expected"
    chi_writer(chi2_2x2_a_df, r = 19, c = 9)

if not os.path.exists("./output"):
    os.mkdir("./output")

output.save(f"./output/result_{input}")